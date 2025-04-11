from collections import defaultdict
from datetime import timedelta, date, time
from django.shortcuts import get_object_or_404
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from rest_framework import status
import re
import openpyxl
from openpyxl.styles import Font
from django.db.models import F, Sum, Q
from django.db.models.functions import TruncMonth
from django.utils.encoding import force_str
from django.db import transaction
from django.utils import timezone
import pandas as pd
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.http import HttpResponse
from django.utils.crypto import get_random_string
from django.utils.dateparse import parse_date
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django_celery_beat.utils import now
from jwt.utils import force_bytes
import logging
from presence_plus.models import User  
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status, permissions, generics, viewsets, mixins,filters
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.request import Request
from django.contrib.auth.models import User
from django.conf import settings
from presence_plus.models import *
from presence_plus.serializers import *
from presence_plus.tasks import *
import uuid
from rest_framework_simplejwt.authentication import JWTAuthentication


User = get_user_model()

class CreateUserView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        emp_num = request.data.get("emp_num")
        name = request.data.get("name")
        department = request.data.get("department") 
        designation_id = request.data.get("designation")
        community_id = request.data.get("community")
        email = request.data.get("email")  
        hire_date = request.data.get("hire_date")
        password = request.data.get("password")  
        username = email  

        current_role = getattr(request.user, "role", "").lower() if request.user else None

        if current_role == "admin":
            new_user_role = "hr"
        elif current_role == "hr":
            new_user_role = "employee"
        else:
            return Response({"error": "You do not have permission to create users"}, status=status.HTTP_403_FORBIDDEN)

        if not all([email, department, name, emp_num, hire_date, designation_id, community_id]):
            return Response({"error": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)

        # if designation_id == 'none' or community_id == 'none':
        #     return Response({"error": "Please select valid Designation and Community"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            designation_id = int(designation_id)
            community_id = int(community_id)
        except ValueError:
            return Response({"error": "Designation and Community must be valid integers."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            hire_date = now().strptime(hire_date, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid hire_date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email).exists():
            return Response({"error": "User with this email already exists"}, status=status.HTTP_400_BAD_REQUEST)

        if Employee.objects.filter(emp_num=emp_num).exists():
            return Response({"error": "Employee Number already exists"}, status=status.HTTP_400_BAD_REQUEST)

        designation_obj = get_object_or_404(Designation, id=designation_id)
        community_obj = get_object_or_404(Community, id=community_id)

        if not password:
            password = get_random_string(length=10)

        try:
            with transaction.atomic():
                user = User.objects.create(
                    email=email,
                    password=make_password(password),
                    role=new_user_role,
                    department=department,
                    username=username
                )

                Employee.objects.create(
                    user=user,
                    name=name,
                    emp_num=emp_num,
                    hire_date=hire_date,
                    designation=designation_obj,
                    community=community_obj
                )

                subject = "Your Account Has Been Created"
                message = f"""
                Hello {name},

                Your account has been successfully created.

                Login Credentials:
                - Email: {email}
                - Password: {password}

                Please log in and change your password immediately.

                Regards,
                Team Presence
                """
                send_mail(subject, message, settings.EMAIL_HOST_USER, [email])

            return Response({"message": f"{new_user_role.capitalize()} user created successfully. Credentials sent to email."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#######     login   ##########################
User = get_user_model()

class LoginView(APIView):
    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({"error": "Invalid email or password"}, status=status.HTTP_401_UNAUTHORIZED)

        if not check_password(password, user.password):
            return Response({"error": "Invalid email or password"}, status=status.HTTP_401_UNAUTHORIZED)

        refresh = RefreshToken.for_user(user)

        role = user.role.lower()
        dashboard_url = {
            "admin": "/admin-dashboard",
            "hr": "/hr-dashboard",
            "employee": "/emp-dashboard"
        }.get(role, None)

        if not dashboard_url:
            return Response({"error": "Invalid role"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "access": str(refresh.access_token),  #  Fix key from 'token' to 'access'
            "refresh": str(refresh),
            "role": user.role,
            "redirect_url": dashboard_url,
        }, status=status.HTTP_200_OK)


################        forget password #####################
User = get_user_model()


class ForgotPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        # Get the email from the request data
        email = request.data.get('email')
        if not email:
            return Response({"error": "Email is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Check if a user with the provided email exists
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            # Return a success message even if the user doesn't exist (to avoid leaking information)
            return Response({
                "message": "If an account with this email exists, you will receive a password reset email shortly."
            }, status=status.HTTP_200_OK)

        # Generate a password reset token and encode the user's primary key
        token_generator = PasswordResetTokenGenerator()
        token = token_generator.make_token(user)
        uidb64 = urlsafe_base64_encode(force_bytes(str(user.pk)))

        # Build the password reset URL
        # reset_url = f"{request.build_absolute_uri('/api/reset-password/')}?uid={uidb64}&token={token}"

        frontend_url = "http://192.168.251.97:5000"
        reset_url = f"{frontend_url}/reset-password/{uidb64}/{token}"

        # Prepare the email content
        subject = "Password Reset Request"
        message = (
            f"Hi {user.username},\n\n"
            "We received a request to reset your password. Click the link below to set a new password:\n"
            f"{reset_url}\n\n"
            "If you didn't request this, please ignore this email.\n"
            "Thank you."
        )

        try:
            # Debugging: Print email details
            # print(f"Attempting to send email to {user.email}...")  # Debugging
            # print(f"Reset URL: {reset_url}")  # Debugging
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            print("Email sent successfully!")  # Debugging
        except Exception as e:
            # Log the error and return a 500 response
            print(f"Error sending email: {e}")  # Debugging
            return Response({"error": "Failed to send email."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Return a success response
        return Response({
            "message": "If an account with this email exists, you will receive a password reset email shortly."
        }, status=status.HTTP_200_OK)

######################      reset password  #####################33
User = get_user_model()


class ResetPasswordView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PasswordResetSerializer(data=request.data)
        if serializer.is_valid():
            uidb64 = serializer.validated_data.get('uid')
            token = serializer.validated_data.get('token')
            new_password = serializer.validated_data.get('new_password')

            try:
                # Decode the uidb64 to get the user id
                uid = force_str(urlsafe_base64_decode(uidb64))
                user = User.objects.get(pk=uid)
            except (TypeError, ValueError, OverflowError, User.DoesNotExist):
                return Response({"error": "Invalid uid."}, status=status.HTTP_400_BAD_REQUEST)

            # Validate the token
            token_generator = PasswordResetTokenGenerator()
            if not token_generator.check_token(user, token):
                return Response({"error": "Invalid or expired token."}, status=status.HTTP_400_BAD_REQUEST)

            # Set the new password and save the user
            user.set_password(new_password)
            user.save()

            return Response({"message": "Password reset successful."}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#################       editing role    #########################

class EditUserRoleView(APIView):
    def post(self, request, user_id):
        # Ensure that the currently logged-in user is either admin or hr
        current_user = request.user
        if current_user.role.lower() not in ["admin", "hr"]:
            return Response({"error": "Only admin and HR can edit user roles."}, status=status.HTTP_403_FORBIDDEN)

        # Get the user to update by user_id
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        new_role = request.data.get("role")
        if not new_role:
            return Response({"error": "Role is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Ensure the role is one of the allowed roles
        if new_role.lower() not in ["hr", "employee"]:
            return Response({"error": "Invalid role"}, status=status.HTTP_400_BAD_REQUEST)

        # If the current user is admin, they can assign any role (HR or Employee)
        # If the current user is HR, they can only assign "employee" role
        if current_user.role.lower() == "hr" and new_role.lower() != "employee":
            return Response({"error": "HR can only assign the 'employee' role."}, status=status.HTTP_403_FORBIDDEN)

        # Update the user's role
        user.role = new_role.lower()
        user.save()

        return Response({"message": f"User role updated to {new_role}"}, status=status.HTTP_200_OK)

############    hr view #################

User = get_user_model()

class HRListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch HR users and join with Employee table to get the name
        hr_users = User.objects.filter(role__iexact="hr").select_related("employee").values(
            "id", "email", "role", "department", "employee__name"
        )

        if not hr_users:
            return Response({"message": "No HR users found."}, status=status.HTTP_404_NOT_FOUND)

        return Response({"hr_users": list(hr_users)}, status=status.HTTP_200_OK)

#############       delete      ######################

class Delete(APIView):
    permission_classes = [IsAuthenticated]  # Only authenticated users can delete HR

    def delete(self, request, hr_id):
        user = request.user  # Get the currently logged-in user

        # Ensure only Admins can delete HR users
        if user.role.lower() != "admin":
            return Response({"error": "Only Admins can delete HR users"}, status=status.HTTP_403_FORBIDDEN)

        try:
            hr_user = User.objects.get(id=hr_id, role__iexact="hr")
            hr_user.delete()
            return Response({"message": "HR user deleted successfully."}, status=status.HTTP_200_OK)
        except User.DoesNotExist:
            return Response({"error": "HR user not found."}, status=status.HTTP_404_NOT_FOUND)

#################   admin profile   #############################

User = get_user_model()

class AdminProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = AdminProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        return self.request.user

#################       dashboard admin view       ########################

User = get_user_model()

class DashboardCountsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Count by role from User model
        total_employees = User.objects.filter(role__iexact="Employee").count()
        total_hr = User.objects.filter(role__iexact="HR").count()

        # Get all Employee objects whose user is HR
        hr_employees = Employee.objects.filter(user__role__iexact="HR")

        # Count leave requests for HRs only
        leave_requests = LeaveRequest.objects.filter(
            status__iexact="pending",
            employee__in=hr_employees
        ).count()

        leave_cancellations = LeaveRequest.objects.filter(
            status__iexact="Cancellation Pending",
            employee__in=hr_employees
        ).count()

        return Response({
            "total_employees": total_employees,
            "total_hr": total_hr,
            "leave_requests": leave_requests,
            "leave_cancellations": leave_cancellations
        }, status=status.HTTP_200_OK)

##################      admin dashboard statistics  ###############
import calendar
class AttendanceStatsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        current_year = datetime.now().year
        attendance_summary = []

        for month in range(1, 13):
            first_day = date(current_year, month, 1)
            last_day = date(current_year, month, calendar.monthrange(current_year, month)[1])

            monthly_attendance = Attendance.objects.filter(date__range=(first_day, last_day))
            total_count = monthly_attendance.count()

            present_count = monthly_attendance.filter(status="present").count()
            absent_count = monthly_attendance.filter(status="absent").count()
            late_count = monthly_attendance.filter(status="late").count()

            attendance_summary.append({
                "month": first_day.strftime("%B"),
                "present_percentage": round((present_count / total_count) * 100, 2) if total_count else 0,
                "absent_percentage": round((absent_count / total_count) * 100, 2) if total_count else 0,
                "late_percentage": round((late_count / total_count) * 100, 2) if total_count else 0
            })

        return Response({
            "year": current_year,
            "attendance_stats": attendance_summary
        }, status=200)

###############     Leave request view  ##################
class LeaveRequestListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Retrieve all pending leave requests submitted by employees, accessible only by HR."""
        try:
            if request.user.role.lower() != "admin":
                return Response(
                    {"error": "Access denied! Only Admin can manage pending leave requests."},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Fetch leave requests that are pending and submitted by employees
            pending_leaves = LeaveRequest.objects.filter(
                employee__user__role="hr", status="Pending"
            )

            # Serialize the data
            serializer = LeaveRequestSerializer(pending_leaves, many=True)

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#################  leave policy creation #############

class LeavePolicyCreateView(viewsets.ModelViewSet):
    queryset = LeavePolicy.objects.all()
    serializer_class = LeavePolicySerializer

    def destroy(self, request, *args, **kwargs):
        """Mark policy as inactive instead of deleting"""
        instance = self.get_object()
        instance.status = "inactive"
        instance.save()
        return Response({"message": "Policy marked as inactive"}, status=status.HTTP_200_OK)

########### policy view ################

class LeavePolicyViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing leave policies.
    """
    queryset = LeavePolicy.objects.all().order_by('-id')  # Fetch all policies, ordered by latest
    serializer_class = LeavePolicySerializer

    def list(self, request, *args, **kwargs):
        """
        GET method: Retrieve all leave policies.
        """
        policies = self.get_queryset()
        serializer = self.get_serializer(policies, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk=None):
        """
        GET method: Retrieve a single leave policy by ID.
        """
        try:
            policy = LeavePolicy.objects.get(pk=pk)
            serializer = self.get_serializer(policy)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except LeavePolicy.DoesNotExist:
            return Response({"error": "Leave policy not found"}, status=status.HTTP_404_NOT_FOUND)

############    leave policy updation   #############

class LeavePolicyDetailView(viewsets.GenericViewSet,
                         mixins.RetrieveModelMixin,
                         mixins.UpdateModelMixin,
                         mixins.DestroyModelMixin):
    """
    ViewSet for handling:
    - GET: Retrieve a single leave policy
    - PUT/PATCH: Update a leave policy
    - DELETE: Delete a leave policy
    """
    queryset = LeavePolicy.objects.all()
    serializer_class = LeavePolicyUpdateSerializer

########### work policy creation and updation   #############

class WorkTimePolicyCreateListView(viewsets.ModelViewSet):
    """
    Handles:
    - GET: List all work time policies
    - POST: Create a new work time policy
    """
    queryset = WorkingHours.objects.all()
    serializer_class = WorkTimePolicySerializer


class WorkTimePolicyDetailView(viewsets.ModelViewSet):
    """
    Handles:
    - GET: Retrieve a specific work time policy
    - PUT/PATCH: Update a work time policy
    - DELETE: Delete a work time policy
    """
    queryset = WorkingHours.objects.all()
    serializer_class = WorkTimePolicySerializer

########### Holiday policy creation & updation  ############

class PublicHolidayView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = PublicHolidaySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        holidays = PublicHoliday.objects.all().order_by('date')  
        serializer = PublicHolidaySerializer(holidays, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class LeaveTypeCreateView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                         
        """Create a new public holiday leave type"""
        serializer = LeaveTypeSerializer(data=request.data)                                                     
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Public holiday leave type created successfully!", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request):
        """Get all leave types"""
        leave_types = LeaveType.objects.all()
        serializer = LeaveTypeSerializer(leave_types, many=True)
        return Response(
            {
                "message": "Leave types retrieved successfully!",
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

#################   logout  ##############
class LogoutView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()  #  Blacklist the token to invalidate it
                return Response({"message": "Successfully logged out"}, status=200)
            return Response({"error": "Refresh token is required"}, status=400)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

#############   Employee dashboard  #############

class EmployeeDashboardView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        employee = request.user.employee
        today = date.today()
        first_day_of_month = today.replace(day=1)

        # Fetch today's attendance
        attendance = Attendance.objects.filter(employee=employee, date=today).first()
        check_in = attendance.check_in if attendance else None
        check_out = attendance.check_out if attendance else None
        late_status = attendance.status == "late" if attendance else False

        # # Fetch today's shift
        # shift_assignment = EmployeeShiftAssignment.objects.select_related('shift').filter(
        #     employee=employee, date=today
        # ).first()
        # shift = shift_assignment.shift.shift_type if shift_assignment else "Not Assigned"

        # Fetch overtime data
        overtime_today = Overtime.objects.filter(employee=employee, date=today).aggregate(Sum("hours"))["hours__sum"] or 0
        total_overtime = Overtime.objects.filter(employee=employee).aggregate(Sum("hours"))["hours__sum"] or 0

        # Calculate monthly attendance percentage
        total_working_days = Attendance.objects.filter(employee=employee, date__gte=first_day_of_month).count()
        present_days = Attendance.objects.filter(employee=employee, date__gte=first_day_of_month, status="present").count()
        attendance_percentage = round((present_days / total_working_days * 100), 2) if total_working_days else 0

        #  Fetch daily attendance for graphical representation
        attendance_graph_data = []
        for day in range(1, today.day + 1):
            date_obj = first_day_of_month.replace(day=day)
            status = Attendance.objects.filter(employee=employee, date=date_obj).first()
            status_value = 1 if status and status.status == "present" else 0  # Present = 1, Absent = 0
            attendance_graph_data.append({"date": date_obj.strftime("%d %b"), "status": status_value})

        return Response({
            "check_in": check_in.strftime("%H:%M:%S") if check_in else "Not Available",
            "check_out": check_out.strftime("%H:%M:%S") if check_out else "Not Available",
            "late": late_status,
            "overtime_today": f"{overtime_today} hours",
            "total_overtime": f"{total_overtime} hours",
            "attendance_percentage": attendance_percentage,
            # "shift": shift,
            "date": today.strftime("%d %b"),
            "attendance_graph_data": attendance_graph_data  # Graph Data
        })

#################   employee leave request and history view ##########################

logger = logging.getLogger(__name__)  #  Proper Logging Setup

class LeaveRequestView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            employee = request.user.employee
            start_date = request.data.get("start_date")
            end_date = request.data.get("end_date")
            leave_type_id = request.data.get("leave_type")  # Should be LeavePolicy ID
            reason = request.data.get("reason")
            image = request.FILES.get("image")  

            if not start_date or not end_date or not leave_type_id or not reason:
                return Response({"error": "All fields are required!"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                leave_type_id = int(leave_type_id)
            except ValueError:
                return Response({"error": "leave_type must be an integer!"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = date.fromisoformat(start_date)
                end_date = date.fromisoformat(end_date)
            except ValueError:
                return Response({"error": "Invalid date format! Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

            if start_date > end_date:
                return Response({"error": "start_date cannot be after end_date!"}, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee already has a leave request on this date
            overlapping_requests = LeaveRequest.objects.filter(
                employee=employee,
                start_date__lte=end_date,
                end_date__gte=start_date
            ).exists()

            if overlapping_requests:
                return Response({"error": "You already have a leave request for this date range!"}, status=status.HTTP_400_BAD_REQUEST)

            # Fetch Leave Policy (Instead of LeaveType)
            leave_policy = get_object_or_404(LeavePolicy, id=leave_type_id)
            leave_type = leave_policy.leave_type  #  Extract the leave type

            #  Fetch Leave Balance
            leave_balance = LeaveBalance.objects.filter(employee=employee, leave_policy=leave_policy).first()
            if not leave_balance:
                return Response({"error": "Leave balance not found!"}, status=status.HTTP_404_NOT_FOUND)

            available_days = max((leave_balance.total or 0) - (leave_balance.used or 0), 0)
            requested_days = (end_date - start_date).days + 1

            if requested_days > available_days:
                return Response({"error": "Insufficient leave balance!"}, status=status.HTTP_400_BAD_REQUEST)

            #  Create Leave Request
            leave_request = LeaveRequest.objects.create(
                employee=employee,
                start_date=start_date,
                end_date=end_date,
                status="Pending",
                reason=reason,
                leave_policy=leave_policy,
                image=image,
            )

            #  Update Leave Balance
            leave_balance.used += requested_days
            leave_balance.save()

            #  Create Leave Transaction
            LeaveTransaction.objects.create(
                employee=employee,
                transaction_type="Leave Request",
                date=date.today(),
                debit=requested_days,
                pending=True,
                leave_policy=leave_policy
            )

            return Response({"message": "Leave request submitted successfully!"}, status=status.HTTP_201_CREATED)

        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request):
        """Get all leave requests for the authenticated employee."""
        try:
            employee = request.user.employee
            leave_requests = LeaveRequest.objects.filter(employee=employee).order_by("-start_date")

            if not leave_requests.exists():
                return Response({"message": "No leave requests found."}, status=status.HTTP_200_OK)

            data = [
                {
                    "id": leave.id,
                    "name": leave.employee.name,
                    "start_date": leave.start_date,
                    "end_date": leave.end_date,
                    "leave_type": leave.leave_policy.leave_type,  # Fixed here
                    "status": leave.status,
                    "reason": leave.reason,
                    "cancellation_reason" : leave.cancellation_reason,
                    "image" : request.build_absolute_uri(leave.image.url) if leave.image else None,
                    
                }
                for leave in leave_requests
            ]

            return Response(data, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class LeaveTypeListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    logger = logging.getLogger(__name__)  

    def get(self, request):
        try:
            logger.info("🚀 API Request Received!")

            employee = Employee.objects.get(user=request.user)
            logger.info(f"Authenticated Employee: {employee}")

            #  Fetch leave balances
            leave_balances = LeaveBalance.objects.filter(employee=employee).annotate(
                balance=F("total") - F("used")
            ).filter(balance__gt=0)

            logger.info(f"Leave Balances Retrieved: {list(leave_balances.values('id', 'total', 'used', 'balance'))}")

            # Extract leave policies
            leave_policies = [lb.leave_policy for lb in leave_balances if lb.balance > 0]
            logger.info(f"Leave Policies Extracted: {leave_policies}")

            #  Prepare response data
            data = [{"id": lp.id, "name": lp.leave_type} for lp in leave_policies]
            logger.info(f"Response Data: {data}")

            return Response(data, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            logger.error(f"Employee not found for user {request.user}")
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Unexpected error in LeaveTypeListView: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

###########     leave balance view  ##############
class LeaveBalanceSummaryView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Retrieve total available and used leave for the logged-in employee."""
        try:
            employee = getattr(request.user, 'employee', None)
            if not employee:
                return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)

            leave_balances = LeaveBalance.objects.filter(employee=employee)

            # Total and available leave from balance model
            total_leave = sum(lb.total or 0 for lb in leave_balances)
            used_leave_from_balance = sum(lb.used or 0 for lb in leave_balances)
            available_leave = total_leave - used_leave_from_balance

            # Get leave requests by status
            approved_leaves = LeaveRequest.objects.filter(
                employee=employee,
                status__in=["Accepted", "Approved"]
            )
            rejected_leaves = LeaveRequest.objects.filter(employee=employee, status="Rejected")
            canceled_leaves = LeaveRequest.objects.filter(employee=employee, status="Cancelled")

            def get_leave_days(lr):
                if lr.start_date and lr.end_date:
                    return (lr.end_date - lr.start_date).days + 1
                return 0

            approved_leave_days = sum(get_leave_days(lr) for lr in approved_leaves)
            rejected_leave_days = sum(get_leave_days(lr) for lr in rejected_leaves)
            refunded_leave_days = sum(get_leave_days(lr) for lr in canceled_leaves)

            # Adjust available leave based on canceled and rejected requests
            adjusted_available_leave = available_leave + refunded_leave_days + rejected_leave_days

            data = {
                "total_leave": total_leave,
                "used_leave": approved_leave_days,
                "refunded_leave": refunded_leave_days,
                "rejected_leave": rejected_leave_days,
                "available_leave": adjusted_available_leave,
                "overall_summary": {
                    "total_leaves_used": approved_leave_days - refunded_leave_days,
                    "total_remaining_leaves": adjusted_available_leave,
                },
            }

            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def patch(self, request, pk):
        try:
            leave_request = LeaveRequest.objects.get(pk=pk)
            new_status = request.data.get('status')

            # Only allow HR to update
            if request.user.role != 'HR':
                return Response({"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN)

            old_status = leave_request.status
            leave_request.status = new_status
            leave_request.save()

            # Refund logic if status changes to Rejected or Cancelled
            if new_status in ['Rejected', 'Cancelled'] and old_status == 'Approved':
                leave_days = (leave_request.end_date - leave_request.start_date).days + 1

                leave_balance = LeaveBalance.objects.get(
                    employee=leave_request.employee,
                    leave_type=leave_request.leave_type
                )

                leave_balance.used = max(0, leave_balance.used - leave_days)
                leave_balance.save()

            return Response({"message": f"Leave status updated to {new_status}"}, status=status.HTTP_200_OK)

        except LeaveRequest.DoesNotExist:
            return Response({"error": "Leave request not found"}, status=status.HTTP_404_NOT_FOUND)

        except LeaveBalance.DoesNotExist:
            return Response({"error": "Leave balance not found"}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#############   Leave cancellation required ####################

class LeaveCancellationRequestView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def put(self, request, leave_id):
        """Employee requests leave cancellation (Needs HR approval)"""
        try:
            leave_request = LeaveRequest.objects.get(
                id=leave_id, employee=request.user.employee, status="Approved"
            )
        except LeaveRequest.DoesNotExist:
            return Response(
                {"error": "Leave request not found or cannot be cancelled!"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get the cancellation reason from request data
        cancellation_reason = request.data.get("cancellation_reason", "").strip()
        if not cancellation_reason:
            return Response({"error": "Cancellation reason is required!"}, status=status.HTTP_400_BAD_REQUEST)

        # Update leave request with cancellation details
        leave_request.cancellation_request = True
        leave_request.status = "Cancellation Pending"
        leave_request.cancellation_reason = cancellation_reason
        leave_request.save()

        # Create leave transaction entry
        LeaveTransaction.objects.create(
            employee=request.user.employee,
            transaction_type="Cancellation Request",
            date=date.today(),
            pending=True,
            leave_policy=leave_request.leave_policy  # Ensure leave_policy is assigned
        )

        return Response(
            {"message": "Leave cancellation request submitted successfully!"},
            status=status.HTTP_200_OK
        ) 
    
###############     HR leave cancellation view  #################################

class LeaveCancellationApprovalView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def put(self, request, leave_id):
        """HR approves or rejects leave cancellation"""

        if request.user.role.lower() != "hr":
            return Response({"error": "Unauthorized action!"}, status=status.HTTP_403_FORBIDDEN)

        decision = request.data.get("decision")  # Accept or Reject

        # Add this validation
        if not decision:
            return Response({"error": "Missing 'decision' in request body!"}, status=status.HTTP_400_BAD_REQUEST)

        decision = decision.lower()

        try:
            leave_request = LeaveRequest.objects.get(id=leave_id, status="Cancellation Pending")
        except LeaveRequest.DoesNotExist:
            return Response({"error": "Cancellation request not found!"}, status=status.HTTP_404_NOT_FOUND)

        if decision == "approve":
            leave_request.status = "Cancelled"
            leave_request.save()

            # Refund leave balance
            requested_days = (leave_request.end_date - leave_request.start_date).days + 1
            leave_balance = LeaveBalance.objects.filter(employee=leave_request.employee).first()
            if leave_balance:
                leave_balance.total += requested_days  # Increase total leave
                leave_balance.used -= requested_days  #  Reduce used leave 
                leave_balance.save()

            # Update transaction as completed
            LeaveTransaction.objects.create(
                employee=leave_request.employee,
                transaction_type="Leave Cancellation Approved",
                date=date.today(),
                credit=requested_days,
                leave_policy=leave_request.leave_policy  
            )

            return Response({"message": "Leave cancellation approved successfully!"}, status=status.HTTP_200_OK)

        elif decision == "reject":
            leave_request.status = "Cancellation Rejected"  # Revert back to approved leave
            leave_request.cancellation_request = False
            leave_request.save()

            # Mark transaction as rejected
            LeaveTransaction.objects.create(
                employee=leave_request.employee,
                transaction_type="Leave Cancellation Rejected",
                date=date.today(),
                pending=False,
                leave_policy=leave_request.leave_policy  
            )

            return Response({"message": "Leave cancellation rejected!"}, status=status.HTTP_200_OK)

        return Response({"error": "Invalid decision value!"}, status=status.HTTP_400_BAD_REQUEST)

    
################    employee overtime view  #####################
from django.db.models.functions import TruncMonth
class OvertimeStatsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Retrieve overtime stats grouped by month"""
        employee = request.user.employee

        # Group by month and calculate total overtime hours per month
        overtime_data = (
            Overtime.objects.filter(employee=employee, status="completed")
            .annotate(month=TruncMonth("date"))  
            .values("month")
            .annotate(total_hours=Sum("hours"))
            .order_by("month")
        )

        # Convert the data into a more readable format (Month Name → Hours)
        monthly_overtime = {entry["month"].strftime("%b"): entry["total_hours"] for entry in overtime_data}

        # Get total overtime hours
        total_hours = sum(monthly_overtime.values())

        # Get current month stats
        current_month = datetime.now().strftime("%b")
        selected_month_hours = monthly_overtime.get(current_month, 0)

        return Response(
            {
                "monthly_overtime": monthly_overtime,
                "total_hours": total_hours,
                "selected_month": {"month": current_month, "hours": selected_month_hours},
            },
            status=status.HTTP_200_OK,
        )

class OvertimeAssignmentView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Retrieve upcoming, missed, and completed overtime assignments"""
        try:
            # Get the employee object
            employee = getattr(request.user, "employee", None)
            if not employee:
                return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)

            print(f"Employee ID: {employee.id}")  # Debugging line

            # Fetch the employee's latest attendance record
            latest_attendance = Attendance.objects.filter(employee=employee).order_by("-date").first()

            if latest_attendance:
                print(f"Latest Check-out Time: {latest_attendance.check_out}")  # Debugging line

                # Ensure check_out is a datetime object
                latest_checkout_datetime = datetime.combine(latest_attendance.date, latest_attendance.check_out)

                # Get all overtime entries assigned to this employee
                overtime_entries = Overtime.objects.filter(employee=employee, status__in=["upcoming", "missed"])

                for overtime in overtime_entries:
                    overtime_end_time = datetime.combine(overtime.date, time(0, 0)) + timedelta(hours=overtime.hours)
                    today = datetime.now().date()

                    if overtime.date > today:
                    # Future overtime should remain "upcoming"
                        overtime.status = "upcoming"
                    elif latest_checkout_datetime >= overtime_end_time:
                    # Overtime is completed
                        overtime.status = "completed"
                    else:
                    # If overtime date has passed and was not completed, it's missed
                        overtime.status = "missed"

                    overtime.save() 

            # Retrieve updated overtime records
            upcoming_overtime = list(Overtime.objects.filter(employee=employee, status="upcoming").values("date", "hours", "reason"))
            missed_overtime = list(Overtime.objects.filter(employee=employee, status="missed").values("date", "hours", "reason"))
            completed_overtime = list(Overtime.objects.filter(employee=employee, status="completed").values("date", "hours", "reason"))

            print("Upcoming Overtime:", upcoming_overtime)  # Debugging line
            print("Missed Overtime:", missed_overtime)  # Debugging line
            print("Completed Overtime:", completed_overtime)  # Debugging line

            return Response(
                {
                    "upcoming_overtime": upcoming_overtime,
                    "missed_overtime": missed_overtime,
                    "completed_overtime": completed_overtime,
                },
                status=status.HTTP_200_OK,
            )

        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)
############    employee attendance stat    ##################
class MonthlyAttendanceStatisticsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        employee = request.user.employee

        # Get month and year from frontend request (default to current month & year)
        month = request.GET.get("month")
        year = request.GET.get("year")

        if not month or not year:
            return Response({"error": "Month and year are required."}, status=400)

        month = int(month)
        year = int(year)

        # Calculate first and last day of the selected month
        first_day_of_month = date(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        # Count attendance records for the selected month
        attendance_counts = Attendance.objects.filter(
            employee=employee,
            date__range=(first_day_of_month, last_day_of_month)
        ).values("status").annotate(count=Count("status"))

        # Initialize attendance counts
        total_days = (last_day_of_month - first_day_of_month).days + 1
        present_count = 0
        late_count = 0
        absent_count = 0

        # Populate attendance counts based on fetched records
        for record in attendance_counts:
            if record["status"] == "present":
                present_count = record["count"]
            elif record["status"] == "late":
                late_count = record["count"]
            elif record["status"] == "absent":
                absent_count = record["count"]

        # Calculate percentages (avoid division by zero)
        if total_days > 0:
            present_percentage = round((present_count / total_days) * 100, 2)
            late_percentage = round((late_count / total_days) * 100, 2)
            absent_percentage = round((absent_count / total_days) * 100, 2)
        else:
            present_percentage = late_percentage = absent_percentage = 0

        return Response({
            "month": first_day_of_month.strftime("%B"),
            "year": year,
            "total_days": total_days,
            "attendance_statistics": {
                "present": {"count": present_count, "percentage": present_percentage},
                "late": {"count": late_count, "percentage": late_percentage},
                "absent": {"count": absent_count, "percentage": absent_percentage},
            },
        }, status=200)

###########     employee attendance manual request and view ##################

class AttendanceListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Fetch attendance records and manual attendance requests for the logged-in employee"""
        try:
            employee = request.user.employee

            # Fetch attendance records
            #attendance_records = Attendance.objects.filter(employee=employee)

            # Fetch manual attendance requests
            attendance_requests = AttendanceRequest.objects.filter(employee=employee)

            # Serialize Attendance data
            # attendance_data = [
            #     {
            #         "id": att.id,
            #         "date": att.date,
            #         "check_in": att.check_in,
            #         "check_out": att.check_out,
            #         "status": "Approved",  # Since Attendance records are already approved
            #         "work_type":  getattr(att, "work_type", None),
            #         "location": getattr(att, "location", None),
            #         "image": request.build_absolute_uri(att.image.url) if hasattr(att, "image") and att.image else None,
            #     }
            #     for att in attendance_records
            # ]

            # Serialize Attendance Request data
            request_data = [
                {
                    "id": req.id,
                    "date": req.date,
                    "check_in": req.check_in,
                    "check_out": req.check_out,
                    "status": req.status,  # Pending, Approved, or Rejected
                    "work_type":  getattr(req, "work_type", None),
                    "location": getattr(req, "location", None),
                    "image": request.build_absolute_uri(req.image.url) if hasattr(req, "image") and req.image else None,
                }
                for req in attendance_requests
            ]

            # Combine both lists
            response_data =  request_data

            return Response(response_data, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=status.HTTP_404_NOT_FOUND)

class AttendanceRequestView(generics.ListCreateAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = AttendanceRequestSerializer

    def get_queryset(self):
        return AttendanceRequest.objects.filter(employee=self.request.user.employee)

    def perform_create(self, serializer):
        employee = self.request.user.employee
        requested_date = serializer.validated_data.get('date', timezone.now().date())

        # Rule 1: Check if attendance already exists for this date
        if Attendance.objects.filter(employee=employee, date=requested_date).exists():
            raise ValidationError(
                {"error": f"Attendance already recorded for {requested_date}. Request denied."},
                code="attendance_exists"
            )

        # Rule 2: Check if an attendance request already exists for this date
        if AttendanceRequest.objects.filter(employee=employee, date=requested_date).exists():
            raise ValidationError(
                {"error": f"You already have a pending attendance request for {requested_date}."},
                code="duplicate_request"
            )

        # Rule 3 (Optional): Restrict requests to the past 30 days
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        if requested_date < thirty_days_ago:
            raise ValidationError(
                {"error": "Cannot request attendance for dates older than 30 days."},
                code="date_too_old"
            )

        # If all checks pass, save the request
        serializer.save(employee=employee, date=requested_date) 

class AttendanceRequestApprovalView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAdminUser]  # Only HR/Admin can approve/reject requests

    def post(self, request, pk):
        try:
            attendance_request = AttendanceRequest.objects.get(pk=pk)
            status = request.data.get("status")

            if status not in ["approved", "rejected","pending"]:
                return Response({"error": "Invalid status"}, status=400)

            attendance_request.status = status
            attendance_request.save()

            if status == "approved":
                Attendance.objects.create(
                    employee=attendance_request.employee,
                    date=attendance_request.date,
                    check_in=attendance_request.check_in,
                    check_out=attendance_request.check_out,
                    status="present"
                )

            return Response({"message": f"Request {status} successfully"})
        except AttendanceRequest.DoesNotExist:
            return Response({"error": "Request not found"}, status=404)

###############     HR shift assignment ##################
class WorkingHoursViewSet(viewsets.ModelViewSet):
    queryset = WorkingHours.objects.all()
    serializer_class = WorkingHoursSerializer
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]

class EmployeeShiftAssignmentViewSet(viewsets.ModelViewSet):
    authentication_classes = [JWTAuthentication]
    queryset = EmployeeShiftAssignment.objects.all()
    serializer_class = EmployeeShiftAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        date = self.request.query_params.get('date', None)
        if date:
            return self.queryset.filter(date=date)
        return self.queryset

    def perform_create(self, serializer):
        serializer.save()

############    Employee shift view ##############

class EmployeeShiftView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]  # Only logged-in employees can view shifts

    def get(self, request):
        date_str = request.GET.get("date")
        
        # Validate and parse the date
        if date_str:
            shift_date = parse_date(date_str)
            if not shift_date:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        else:
            shift_date = date.today()  # Default to today's date if not provided

        try:
            employee = request.user.employee  # Fetch the related Employee instance
        except Employee.DoesNotExist:
            return Response({"error": "Employee record not found"}, status=404)

        shift_assignment = EmployeeShiftAssignment.objects.filter(
            employee=employee, date=shift_date
        ).select_related("shift").first()

        if shift_assignment:
            shift_data = {
                "date": shift_date,
                "shift": shift_assignment.shift.shift_type,
                "start_time": shift_assignment.shift.start_time,
                "end_time": shift_assignment.shift.end_time
            }
            return Response(shift_data, status=200)
#http://0.0.0.0:8000/empshiftview/?date=2025-03-20 pass in this format
        return Response({"message": "No shift assigned for this date"}, status=404)

class ShiftRosterView(APIView):
    permission_classes = [permissions.IsAuthenticated]  # Only HR/Admin can view shift rosters
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        roster_id = request.GET.get("roster_id")
        shifts = EmployeeShiftAssignment.objects.filter(shift_roster_id=roster_id).select_related("employee", "shift")

        if not shifts.exists():
            return Response({"message": "No shifts assigned for this roster"}, status=404)

        data = [
            {
                "employee": shift.employee.name,
                "date": shift.date,
                "shift": shift.shift.shift_type,
                "start_time": shift.shift.start_time,
                "end_time": shift.shift.end_time,
            }
            for shift in shifts
        ]
        return Response(data)
#http://0.0.0.0:8000/empshiftroster/?roster_id=1 pass in this format

############    employee profile    ##################################

class EmployeeProfileView(APIView):
    authentication_classes = [JWTAuthentication]  
    permission_classes = [IsAuthenticated]  

    def get(self, request):
        print("Authenticated user:", request.user)  # Debugging
        if request.user.is_anonymous:
            return Response({"error": "User is not authenticated"}, status=401)

        employee = Employee.objects.select_related("user").get(user=request.user)

        profile_data = {
            "name": employee.name,
            "email": employee.user.email,
            "employee_id": employee.emp_num,
            "hire_date": employee.hire_date,
            "department": employee.user.department,
            "position": employee.designation.desig_name if employee.designation else None,
            "community_name": employee.community.community_name if employee.community else None,
            "image": request.build_absolute_uri(employee.image.url) if employee.image else None
        }

        return Response(profile_data)

    def patch(self, request):
        employee = Employee.objects.get(user=request.user)

        if "image" in request.data:
            employee.image = request.data["image"]
            employee.save()

            return Response({
                "message": "Profile updated successfully",
                "profile_image": request.build_absolute_uri(employee.image.url) if employee.image else None
            })

        return Response({"error": "Only profile image can be updated"}, status=400)

################    change password #####################

class ChangePasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def patch(self, request):
        user = request.user
        current_password = request.data.get("current_password")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not check_password(current_password, user.password):
            return Response({"error": "Current password is incorrect"}, status=400)

        if new_password != confirm_password:
            return Response({"error": "New passwords do not match"}, status=400)

        if not self.validate_password_strength(new_password):
            return Response(
                {
                    "error": "Password must be at least 8 characters long, include a number, an uppercase letter, and a special character."},
                    status=400
                )

        user.set_password(new_password)
        user.save()

        return Response({"message": "Password updated successfully"}, status=200)

    def validate_password_strength(self, password):
        """Ensure password is strong."""
        return (
            len(password) >= 8 and
            re.search(r"\d", password) and  # At least one digit
            re.search(r"[A-Z]", password) and  # At least one uppercase letter
            re.search(r"[@$!%*?&#]", password)  # At least one special character
        )

###############     employee leave balance and history  ################

class EmployeeLeaveBalanceView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        employee = get_object_or_404(Employee, user=request.user)
        leave_balances = LeaveBalance.objects.filter(employee=employee)

        leave_data = []

        for leave in leave_balances:
            used = leave.used
            total = leave.total
            leave_policy = leave.leave_policy
            leave_type = leave_policy.leave_type

            # Handle unlimited/unpaid leave
            total_display = "∞" if total == float("inf") else total

            leave_requests = LeaveRequest.objects.filter(
                employee=employee,
                leave_policy=leave_policy,
                status__in=["Approved", "Cancel Rejected","Accepted"],
                cancellation_request=False
            )

            print(f"\nLeave Type: {leave_type}")
            print(f"Total Leaves Found: {leave_requests.count()}")

            leave_dates = []
            for leave_req in leave_requests:
                print(f"Leave Request: {leave_req.start_date} to {leave_req.end_date}, Status: {leave_req.status}")

                leave_dates.append({
                    "start_date": leave_req.start_date.strftime("%Y-%m-%d"),
                    "end_date": leave_req.end_date.strftime("%Y-%m-%d"),
                    "status": leave_req.status
                })

            leave_data.append({
                "name": leave_type,
                "used": f"{used}/{total_display} Used",
                "dates": leave_dates
            })

        return Response({"leave_balance": leave_data}, status=200)

############    policy view ###############

class PolicyListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch only active leave policies
        leave_policies = LeavePolicy.objects.filter(status="active").values(
            "leave_type", "carry_forward", "amount", "frequency"
        )

        # Fetch only active public holidays (Fix: Change "leave_type" to "leavetype")
        public_holidays = PublicHoliday.objects.filter(status="active").values(
            "name", "date", "leavetype", "community__community_name"
        )

        # Fetch only active working hours
        working_hours = WorkingHours.objects.filter(status="active").values(
            "shift_type", "start_time", "end_time"
        )

        return Response({
            "leave_policies": list(leave_policies),
            "public_holidays": list(public_holidays),  # Updated field name
            "working_hours": list(working_hours),
        }, status=200)

################    HR dashboard    ###############

class HRDashboardView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = now().date()

        # Total employees
        total_employees = Employee.objects.filter(Q(user__role="Employee") | Q(user__role="employee"))
        serialized_employees = EmployeeSerializers(total_employees, many=True).data

        # Employees on leave today
        on_leave_today = LeaveRequest.objects.filter(
            start_date__lte=today, end_date__gte=today, status="approved"
        ).count()

        # Pending leave requests
        hr_employees = Employee.objects.filter(user__role__iexact="employee")

        # Count leave requests for HRs only
        leave_requests = LeaveRequest.objects.filter(
            status__iexact="pending",
            employee__in=hr_employees
        ).count()

        leave_cancellations = LeaveRequest.objects.filter(
            status__iexact="Cancellation Pending",
            employee__in=hr_employees
        ).count()

        attendance_requests = AttendanceRequest.objects.filter(
            status__iexact="pending",
            employee__in=hr_employees
        ).count()
        
        # Attendance statistics for today
        present_today = Attendance.objects.filter(date=today, status="present").count()
        absent_today = Attendance.objects.filter(date=today, status="absent").count()
        late_today = Attendance.objects.filter(date=today, status="late").count()

        # Get check-in and check-out times for present employees
        present_attendance = Attendance.objects.filter(date=today, status="present").select_related("employee")

        attendance_data = []
        for record in present_attendance:
            attendance_data.append({
                "employee": record.employee.user.username,  # Adjust based on your User model
                "check_in": record.check_in.strftime("%H:%M:%S") if record.check_in else "N/A",
                "check_out": record.check_out.strftime("%H:%M:%S") if record.check_out else "N/A",
            })

        # **Weekly Attendance Statistics (Last 7 Days)**
        start_date = today - timedelta(days=6)  # Get last 7 days
        weekly_attendance = Attendance.objects.filter(date__range=[start_date, today]).values("date", "status").annotate(count=Count("id"))

        # Organize data for graph representation
        weekly_stats = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0})
        for entry in weekly_attendance:
            weekly_stats[entry["date"]][entry["status"]] = entry["count"]

        weekly_attendance_data = [
            {
                "date": (start_date + timedelta(days=i)).strftime("%Y-%m-%d"),
                "present": weekly_stats[start_date + timedelta(days=i)]["present"],
                "absent": weekly_stats[start_date + timedelta(days=i)]["absent"],
                "late": weekly_stats[start_date + timedelta(days=i)]["late"],
            }
            for i in range(7)  # Loop through last 7 days
        ]

        return Response({
            "total_employees": serialized_employees,
            "on_leave_today": on_leave_today,
            "leave_requests": leave_requests,  
            "attendance_request": attendance_requests, 
            "leave_cancellations": leave_cancellations,
            "present": present_today,  
            "absent": absent_today,    
            "late": late_today,        
            "attendance_data": attendance_data,  
            "weekly_attendance": weekly_attendance_data,  # Added Weekly Attendance Stats
        }, status=200)
    
##################  HR employee view    ##########################

class EmployeeDetailView(generics.ListAPIView):
    authentication_classes = [JWTAuthentication]
    queryset = Employee.objects.filter(Q(user__role="Employee") | Q(user__role="employee"))
    serializer_class = EmployeeListSerializer
    permission_classes = [IsAuthenticated]

#################   HR employee manual attendance entry ####################

class ManualAttendanceView(generics.CreateAPIView):
    queryset = Attendance.objects.all()
    serializer_class = ManualAttendanceSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def create(self, request, *args, **kwargs):
        """Handles manual attendance creation with duplicate prevention."""
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            employee = serializer.validated_data.get("employee")
            date = serializer.validated_data.get("date")

            # Check if attendance already exists for this employee and date
            if Attendance.objects.filter(employee=employee, date=date).exists():
                return Response(
                    {"error": f"Attendance already exists for {employee.name} on {date}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            serializer.save()
            return Response(
                {"message": "Attendance recorded successfully", "data": serializer.data},
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
#########   hr requests leave for employee    #####################
from rest_framework.parsers import MultiPartParser, FormParser

class HRRequestLeaveView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [permissions.IsAuthenticated]  

    def post(self, request, *args, **kwargs):
        """HR requests leave for an employee (JSON format)."""
        if request.user.role.lower() != "hr":
            return Response({"error": "Only HR can request leave for employees."}, status=status.HTTP_403_FORBIDDEN)

        # Extract JSON data
        data = request.data  
        employee_id = data.get("employee")
        leave_policy_id = data.get("leave_policy")  
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        reason = data.get("reason")
        image = data.get("image")  # Expecting a base64 encoded string or image URL

        if not all([employee_id, leave_policy_id, start_date, end_date, reason]):
            return Response({"error": "All fields except image are required."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate employee
        employee = get_object_or_404(Employee, id=employee_id)

        # Validate leave policy
        leave_policy = get_object_or_404(LeavePolicy, id=leave_policy_id)

        # Create leave request
        leave_request = LeaveRequest.objects.create(
            employee=employee,
            leave_policy=leave_policy,  
            start_date=start_date,
            end_date=end_date,
            reason=reason,
            status="Accepted", 
        )

        # Handle Image (Optional)
        if image:
            leave_request.image = image  
            leave_request.save()

        serializer = LeaveRequestSerializer(leave_request)
        return Response({"message": "Leave request submitted successfully.", "data": serializer.data}, status=status.HTTP_201_CREATED)

class LeaveTypeDropdownView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        employee_id = request.query_params.get("employee_id")  # Get employee_id from query param

        if not employee_id:
            return Response({"error": "Employee ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id)

            # Get available leave types based on the leave balance
            leave_balances = LeaveBalance.objects.filter(employee=employee).annotate(
                balance=F("total") - F("used")
            ).filter(balance__gt=0)

            available_leave_types = [
                {"id": lb.leave_policy.id, "leave_type": lb.leave_policy.leave_type} for lb in leave_balances
            ]

            return Response(available_leave_types, status=status.HTTP_200_OK)

        except Employee.DoesNotExist:
            return Response({"error": "Invalid Employee ID"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

################    HR leave request view    ####################
class ApproveRejectLeaveView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request, leave_id):
        leave_request = get_object_or_404(LeaveRequest, id=leave_id)

        action = request.data.get("action")  # Either "approve" or "reject"
        reject_reason = request.data.get("reject_reason", "")

        if action == "approve":
            leave_request.status = "Approved"
            leave_request.reject_reason = ""
        elif action == "reject":
            leave_request.status = "Rejected"
            leave_request.reject_reason = reject_reason

        leave_request.save()
        return Response({"message": f"Leave {action}d successfully!"}, status=200)
    
class PendingLeaveRequestsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Retrieve all pending leave requests submitted by employees, accessible only by HR."""
        try:
            if request.user.role.lower() != "hr":
                return Response(
                    {"error": "Access denied! Only HR can manage pending leave requests."},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Fetch leave requests that are pending and submitted by employees
            pending_leaves = LeaveRequest.objects.filter(
                employee__user__role="employee", status="Pending"
            )

            # Serialize the data
            serializer = LeaveRequestSerializer(pending_leaves, many=True)

            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

##############      HR leave cancellation view  #####################

class ApproveRejectCancellationView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request, leave_id):
        leave_request = get_object_or_404(LeaveRequest, id=leave_id)

        if not leave_request.cancellation_request:
            return Response({"error": "No cancellation request for this leave."}, status=400)

        action = request.data.get("action")
        if action == "approve":
            leave_request.status = "Cancelled"
            leave_request.save()
            return Response({"message": "Leave cancellation approved."}, status=200)

        elif action == "reject":
            leave_request.status = "Cancel Rejected"
            leave_request.save()
            return Response({"message": "Leave cancellation rejected."}, status=200)

        return Response({"error": "Invalid action."}, status=400)
    
class PendingLeaveCancellationRequestsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Retrieve leave requests pending cancellation approval based on user role."""
        user_role = request.user.role.lower()  # Role is stored in the User model

        if user_role == "hr":
            # HR can see leave cancellation requests from employees
            pending_cancellations = LeaveRequest.objects.filter(
                status="Cancellation Pending",
                employee__user__role="employee"  # Accessing role from User model
            )

        elif user_role == "admin":
            # Admin can see leave cancellation requests from HRs
            pending_cancellations = LeaveRequest.objects.filter(
                status="Cancellation Pending",
                employee__user__role="hr"  # Accessing role from User model
            )

        else:
            return Response(
                {"error": "Access denied! Only HR and Admin can view leave cancellation requests."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Serialize the data
        data = [
            {
                "id": leave.id,
                "name": leave.employee.name,
                "employee_role": leave.employee.user.role,  
                "leave_type": leave.leave_policy.leave_type if leave.leave_policy else "Unknown",
                "reason": leave.reason,
                "start_date": leave.start_date,
                "end_date": leave.end_date,
                "cancellation_reason": leave.cancellation_reason,
            }
            for leave in pending_cancellations
        ]

        return Response(data, status=status.HTTP_200_OK)

##############  HR and admin leave history view  #####################
class LeaveHistoryView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            user = request.user  
            user_role = user.role.lower()  

            if user_role == "hr":
                # HR can view employees' leave history
                leave_history = LeaveRequest.objects.filter(
                    employee__user__role="employee",  
                    status__in=["Approved", "Cancel Rejected", "approved", "Accepted"]
                ).order_by("-start_date")

            elif user_role == "admin":
                # Admin can view HRs' leave history
                leave_history = LeaveRequest.objects.filter(
                    employee__user__role="hr",  
                    status__in=["Approved", "Cancel Rejected", "approved","Accepted"]
                ).order_by("-start_date")

            else:
                return Response({"error": "Unauthorized access"}, status=403)

            serializer = LeaveRequestHistorySerializer(leave_history, many=True)
            return Response(serializer.data, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)

#########   employee list view  ####################

class EmployeeListView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        employees = Employee.objects.filter(user__role='employee').order_by('name') # Fetch all employees sorted by name
        serializer = EmployeesSerializers(employees, many=True)
        return Response(serializer.data, status=200)

############    HR self portal  #####################
from datetime import datetime

class HRMonthlyAttendanceView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # Ensure only HRs can access their own stats
        if not hasattr(user, "employee"):
            return Response({"error": "User is not an employee"}, status=400)

        today = now().date()
        first_day = today.replace(day=1)

        # Fetch HR’s attendance records for the current month
        attendance_records = Attendance.objects.filter(
            employee=user.employee, date__range=[first_day, today]
        ).order_by("date")

        total_days = (today - first_day).days + 1
        present_days = attendance_records.filter(status="present").count()
        absent_days = attendance_records.filter(status="absent").count()
        late_days = attendance_records.filter(status="late").count()

        # Calculate total overtime hours
        total_overtime_hours = 0
        attendance_list = []

        # Get today's attendance record
        today_attendance = Attendance.objects.filter(employee=user.employee, date=today).first()
        
        # Extract today's check-in and check-out as time objects
        today_check_in = today_attendance.check_in if today_attendance and today_attendance.check_in else None
        today_check_out = today_attendance.check_out if today_attendance and today_attendance.check_out else None

        for record in attendance_records:
            check_in = record.check_in if record.check_in else None
            check_out = record.check_out if record.check_out else None

            # Convert check_in and check_out to datetime
            if check_in and check_out:
                check_in_dt = datetime.combine(record.date, check_in)
                check_out_dt = datetime.combine(record.date, check_out)

                work_hours = (check_out_dt - check_in_dt).total_seconds() / 3600  # Convert to hours
                overtime_hours = max(work_hours - 8, 0)  # Assuming 8 working hours
                total_overtime_hours += overtime_hours
            else:
                work_hours = 0
                overtime_hours = 0

            attendance_list.append({
                "status": record.status,
                "check_in": check_in, 
                "check_out": check_out, 
                "overtime_hours": round(overtime_hours, 2)
            })

        # Calculate attendance percentage
        attendance_percentage = (present_days / total_days) * 100 if total_days > 0 else 0

        data = {
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "late_days": late_days,
            "total_overtime_hours": round(total_overtime_hours, 2),
            "attendance_percentage": round(attendance_percentage, 2),
            "check_in": today_check_in,  
            "check_out": today_check_out, 
        }

        return Response(data, status=200)

#############   create community    ############

class CommunityView(APIView):
    authentication_classes = [JWTAuthentication]  
    permission_classes = [permissions.IsAuthenticated]  

    def get(self, request):
        communities = Community.objects.all()
        serializer = CommunitySerializer(communities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = CommunitySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

##########  create designation  ###################

class DesignationListCreateView(generics.ListCreateAPIView):
    queryset = Designation.objects.all()
    serializer_class = DesignationSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication] 

############    Shift management    #####################

class WorkingHoursListCreateView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        """List all working hours shifts."""
        shifts = WorkingHours.objects.filter(status='active')
        serializer = WorkingHoursSerializer(shifts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        """Create a new shift (working hours)."""
        serializer = WorkingHoursSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AssignShiftView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        try:
            print("Received data:", request.data)  # Debug
    
            # Validate presence of all fields
            required_fields = ['date', 'shift_roster', 'employees']
            if not all(field in request.data for field in required_fields):
                return Response(
                    {"error": f"Required fields: {required_fields}"},
                    status=400
                )
    
            date = request.data['date']
            shift_type = request.data['shift_roster']  # Changed from shift_name
            employee_ids = request.data['employees']
        
            # Validate date format
            try:
                datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD"},
                    status=400
                )
        
            # Validate shift exists - CHANGED THIS PART
            try:
                shift = WorkingHours.objects.get(shift_type=shift_type)  # Using shift_type instead of name
            except WorkingHours.DoesNotExist:
                available_shifts = WorkingHours.objects.values_list('shift_type', flat=True).distinct()
                return Response(
                    {
                        "error": f"Shift type '{shift_type}' doesn't exist",
                        "available_shifts": list(available_shifts)
                    },
                    status=400
                )
        
            # Rest of your code remains the same...
            # Validate employees exist
            valid_employees = Employee.objects.filter(id__in=employee_ids)
            if len(valid_employees) != len(employee_ids):
                invalid_ids = set(employee_ids) - set(valid_employees.values_list('id', flat=True))
                return Response(
                    {"error": f"Invalid employee IDs: {invalid_ids}"},
                    status=400
                )
        
            # Check for existing assignments
            conflicts = EmployeeShiftAssignment.objects.filter(
                employee_id__in=employee_ids,
                date=date
            ).values_list('employee_id', flat=True)
        
            if conflicts:
                return Response(
                    {"error": f"Employees already assigned: {list(conflicts)}"},
                    status=400
                )
        
            # Create assignments
            assignments = [
                EmployeeShiftAssignment(
                    date=date,
                    employee_id=emp_id,
                    shift_id=shift.id
                ) for emp_id in employee_ids
            ]
        
            try:
                EmployeeShiftAssignment.objects.bulk_create(assignments)
                return Response(
                    {"message": f"Assigned {len(assignments)} employees to {shift_type} shift"},
                    status=201
                )
            except Exception as e:
                return Response(
                    {"error": str(e)},
                    status=500
                )
        except Exception as e:
            print(f"Unexpected error: {str(e)}")
            return Response(
                {"error": "Internal server error"},
                status=500
            )

    def put(self, request, assignment_id):
        """Update an existing shift assignment"""
        try:
            assignment = EmployeeShiftAssignment.objects.get(id=assignment_id)
        except EmployeeShiftAssignment.DoesNotExist:
            return Response(
                {"error": "Shift assignment not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        new_shift_id = request.data.get('shift')
        if not new_shift_id:
            return Response(
                {"error": "Shift ID is required for update"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if employee already has another assignment on this date
        if EmployeeShiftAssignment.objects.filter(
            employee=assignment.employee,
            date=assignment.date
        ).exclude(id=assignment_id).exists():
            return Response(
                {"error": "Employee already assigned to another shift on this date"},
                status=status.HTTP_400_BAD_REQUEST
            )

        assignment.shift_id = new_shift_id
        assignment.save()
        
        return Response(
            {
                "message": "Shift assignment updated successfully",
                "data": EmployeeShiftAssignmentSerializer(assignment).data
            },
            status=status.HTTP_200_OK
        )

    def delete(self, request, assignment_id):
        """Delete a shift assignment"""
        try:
            assignment = EmployeeShiftAssignment.objects.get(id=assignment_id)
            assignment.delete()
            return Response(
                {"message": "Shift assignment deleted successfully"},
                status=status.HTTP_204_NO_CONTENT
            )
        except EmployeeShiftAssignment.DoesNotExist:
            return Response(
                {"error": "Shift assignment not found"},
                status=status.HTTP_404_NOT_FOUND
            )
class AssignedShiftView(APIView):
    def get(self, request):
        """Get all shift assignments for a specific date"""
        try:
            date = request.query_params.get('date')
            
            # Validate date parameter exists
            if not date:
                return Response(
                    {"error": "Date parameter is required (e.g., ?date=2025-03-15)"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate date format
            try:
                parsed_date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verify database connection
            from django.db import connection
            connection.ensure_connection()
            
            # Get and serialize data
            assignments = EmployeeShiftAssignment.objects.filter(date=parsed_date)
            serializer = EmployeeShiftAssignmentSerializer(assignments, many=True)
            
            return Response(
                {
                    "date": date,
                    "count": len(serializer.data),
                    "assignments": serializer.data
                },
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            print(f"Server Error: {str(e)}")  # Log to console
            return Response(
                {
                    "error": "Internal server error",
                    "details": str(e) if settings.DEBUG else None
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class EmployeeShiftView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        """Get shifts assigned to the authenticated employee."""
        try:
            # Get employee profile
            try:
                employee = request.user.employee
            except AttributeError:
                return Response(
                    {"error": "User has no associated employee profile"},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Get today's shifts
            today = now().date()
            shifts = EmployeeShiftAssignment.objects.filter(
                employee=employee,
                date=today
            ).order_by('date')
            
            # Serialize data
            serializer = EmployeeShiftAssignmentSerializer(shifts, many=True)
            return Response(
                {
                    "date": today,
                    "shifts": serializer.data
                },
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            print(f"Error in EmployeeShiftView: {str(e)}")  # Log the error
            return Response(
                {"error": "Internal server error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ShiftColleaguesView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, date):
        """Get colleagues assigned to the same shift on a given date, excluding the logged-in employee."""
        employee = request.user.employee

        # Get the shift(s) the logged-in employee is assigned to on that date
        shifts = EmployeeShiftAssignment.objects.filter(
            employee=employee, date=date
        ).values_list('shift', flat=True)

        # Get all assignments for the same shift(s) on that date, excluding the logged-in employee
        colleague_assignments = EmployeeShiftAssignment.objects.filter(
            date=date,
            shift__in=shifts
        ).exclude(employee=employee)

        serializer = EmployeeShiftAssignmentSerializer(colleague_assignments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class ShiftRangeView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    # Enhanced color mapping with shift status consideration
    SHIFT_COLORS = {
        'morning': {'active': '#FFEEBA', 'inactive': '#F5F5F5'},
        'intermediate': {'active': '#B8DAFF', 'inactive': '#E0E0E0'},
        'night': {'active': '#D8BFD8', 'inactive': '#EDEDED'},
        'general': {'active': '#C3E6CB', 'inactive': '#DFDFDF'}
    }

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Validate date parameters
        if not start_date or not end_date:
            return Response(
                {"error": "Both start_date and end_date parameters are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get active shifts for the current user in date range
        user_shifts = EmployeeShiftAssignment.objects.filter(
            employee=request.user.employee,
            date__gte=start,
            date__lte=end,
            shift__status='active'  # Only include active shifts
        ).select_related('shift', 'employee')

        # Get active shifts for colleagues (same shifts as user)
        colleague_shifts = EmployeeShiftAssignment.objects.filter(
            date__gte=start,
            date__lte=end,
            shift__in=[s.shift for s in user_shifts],
            shift__status='active'  # Only include active shifts
        ).exclude(employee=request.user.employee).select_related('shift', 'employee')

        # Organize data by date with working hours
        date_map = {}
        for shift_assignment in list(user_shifts) + list(colleague_shifts):
            shift = shift_assignment.shift
            date_str = shift_assignment.date.strftime('%Y-%m-%d')
            
            if date_str not in date_map:
                date_map[date_str] = {
                    'date': date_str,
                    'shifts': {},
                    'color': self.SHIFT_COLORS.get(shift.shift_type, {}).get(shift.status, '#FFFFFF')
                }
            
            if shift.id not in date_map[date_str]['shifts']:
                date_map[date_str]['shifts'][shift.id] = {
                    'shift_id': shift.id,
                    'shift_name': shift.shift_type,
                    'shift_type': shift.shift_type,
                    'status': shift.status,
                    'start_time': shift.start_time.strftime('%H:%M'),
                    'end_time': shift.end_time.strftime('%H:%M'),
                    'employees': []
                }
            
            date_map[date_str]['shifts'][shift.id]['employees'].append({
                'employee_id': shift_assignment.employee.id,
                'name': shift_assignment.employee.user.get_full_name(),
                'is_current_user': shift_assignment.employee == request.user.employee
            })

        # Convert to list format and sort by date
        result = sorted(list(date_map.values()), key=lambda x: x['date'])
        
        return Response({
            'shift_dates': result,
            'color_mapping': self.SHIFT_COLORS,
            'date_range': {
                'start_date': start_date,
                'end_date': end_date
            }
        }, status=status.HTTP_200_OK)
    
class HRShiftRangeView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    # Single color for all assigned dates in HR view
    ASSIGNED_COLOR = '#00008B'  # Blue color for assigned dates

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Validate date parameters
        if not start_date or not end_date:
            return Response(
                {"error": "Both start_date and end_date parameters are required"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get all assigned shifts in date range (HR can see all)
        assigned_shifts = EmployeeShiftAssignment.objects.filter(
            date__gte=start,
            date__lte=end,
            shift__status='active'  # Only include active shifts
        ).select_related('shift', 'employee')

        # Organize data by date - simplified for HR view
        assigned_dates = set()
        shift_details = []
        
        for shift_assignment in assigned_shifts:
            date_str = shift_assignment.date.strftime('%Y-%m-%d')
            assigned_dates.add(date_str)
            
            shift_details.append({
                'date': date_str,
                'shift_id': shift_assignment.shift.id,
                'shift_type': shift_assignment.shift.shift_type,
                'start_time': shift_assignment.shift.start_time.strftime('%H:%M'),
                'end_time': shift_assignment.shift.end_time.strftime('%H:%M'),
                'employee_id': shift_assignment.employee.id,
                'employee_name': shift_assignment.employee.user.get_full_name()
            })

        # Convert to list format and sort by date
        result = sorted([{'date': date, 'color': self.ASSIGNED_COLOR} for date in assigned_dates], 
                       key=lambda x: x['date'])
        
        return Response({
            'assigned_dates': result,
            'shift_details': shift_details,
            'date_range': {
                'start_date': start_date,
                'end_date': end_date
            }
        }, status=status.HTTP_200_OK)
###########     employee dashboard shift view   #############

class EmployeeDasboardShiftView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        """Get today's shift assigned to the authenticated employee."""
        employee = request.user.employee
        today = now().date()

        # Fetch only today's shift for the logged-in employee
        shifts = EmployeeShiftAssignment.objects.filter(employee=employee, date=today)

        # Serialize and return data
        serializer = EmployeeShiftAssignmentSerializer(shifts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

from django.utils.timezone import now

class ShiftColleaguesDashboardView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        """Get colleagues assigned to the same shift today."""
        employee = request.user.employee
        today = now().date()

        # Get the shift of the authenticated employee for today
        employee_shift = EmployeeShiftAssignment.objects.filter(employee=employee, date=today).values_list('shift', flat=True)

        # Get colleagues assigned to the same shift today
        assignments = EmployeeShiftAssignment.objects.filter(date=today, shift__in=employee_shift)

        serializer = EmployeeShiftAssignmentSerializer(assignments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
###########     notifications       #######################

class NotificationListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifications = Notification.objects.filter(user=request.user).order_by('-time_stamp')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

################    Overtime    #####################

class OvertimeSummaryView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = date.today()

        # 1. Calculate total overtime (only 'Completed' status) for all time
        total_overtime = Overtime.objects.filter(
            status__in=["completed", "upcoming"]  # Include both statuses
        ).aggregate(total_hours=Sum('hours'))['total_hours'] or 0

        # 2. Count number of employees who have completed overtime today
        employees_on_ot_today = Overtime.objects.filter(date=today, status="completed").values('employee').distinct().count()

        # 3. Get list of employees with their total completed overtime, designation, and ID
        employees_overtime = (
            Overtime.objects.filter(status__in=["completed","upcoming"])  # Filter by completed status
            .values('employee__id', 'employee__name', 'employee__designation__desig_name')
            .annotate(total_hours=Sum('hours'))
            .order_by('-total_hours')  # Sort by highest overtime
        )

        # 4. Format the response with employee ID
        employee_data = [
            {
                "employee_id": emp["employee__id"],
                "name": emp["employee__name"],
                "designation": emp["employee__designation__desig_name"],
                "total_overtime": emp["total_hours"]
            }
            for emp in employees_overtime
        ]

        return Response({
            "total_overtime": total_overtime,  # Total completed overtime for all time
            "employees_on_ot_today": employees_on_ot_today,
            "employees_overtime": employee_data
        }, status=200)

import openpyxl
from openpyxl.utils import get_column_letter 

class OvertimeSummaryDownloadView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get query parameters
            start_date_str = request.query_params.get('startDate')
            end_date_str = request.query_params.get('endDate')
            #department = request.query_params.get('department', 'All')

            # Validate required parameters
            if not start_date_str or not end_date_str:
                return Response(
                    {'error': 'Both startDate and endDate are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Convert dates
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Build base query
            queryset = Overtime.objects.filter(
                date__gte=start_date,
                date__lte=end_date
            ).select_related('employee', 'employee__user')

            # Filter by department if specified
            # if department != 'All':
            #     queryset = queryset.filter(employee__user__department=department)

            # Annotate with total hours
            report_data = queryset.values(
                'employee__id',
                'employee__name',
                'employee__user__department',
                'date',
                'status'
            ).annotate(
                total_hours=Sum('hours')
            ).order_by('employee__name', 'date')

            # Create Excel workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Overtime Report"

            # Write headers
            headers = ['Employee ID', 'Name', 'Department', 'Date', 'Hours', 'Status']
            ws.append(headers)

            # Write data rows
            for record in report_data:
                ws.append([
                    record['employee__id'],
                    record['employee__name'],
                    record['employee__user__department'],
                    record['date'].strftime('%Y-%m-%d'),
                    float(record['total_hours']),
                    record['status']
                ])

            # Adjust column width for better readability
            for col_num, column_title in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15

            # Create HTTP response for file download
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="overtime_report.xlsx"'

            wb.save(response)
            return response

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class EmployeeOvertimeDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        try:
            employee = get_object_or_404(
                Employee.objects.select_related("designation", "user"),
                id=employee_id
            )

            completed_overtime_records = Overtime.objects.filter(
                employee=employee
            ).order_by('-date')

            filtered_overtime = []
            total_overtime = 0

            for overtime in completed_overtime_records:
                    assigned_hours = overtime.hours
                    actual_worked_hours = 0
                    overtime_status = "upcoming"  # Default for today/future

                    attendance_record = Attendance.objects.filter(
                        employee=employee, date=overtime.date
                    ).first()

                    # Only evaluate if the overtime date is before today
                    if overtime.date < date.today():
                        if attendance_record and attendance_record.check_in and attendance_record.check_out:
                            actual_worked_seconds = (
                                datetime.combine(overtime.date, attendance_record.check_out) -
                                datetime.combine(overtime.date, attendance_record.check_in)
                            ).seconds
                            actual_worked_hours = actual_worked_seconds / 3600
                            overtime_status = "completed" if actual_worked_hours >= assigned_hours else "missed"
                        else:
                            overtime_status = "missed"
                    elif overtime.date == date.today():
                        overtime_status = "upcoming"

                    if overtime_status == "completed":
                        total_overtime += assigned_hours

                    filtered_overtime.append({
                        "date": overtime.date.strftime("%d %b"),
                        "assigned_hours": assigned_hours,
                        "actual_hours": round(actual_worked_hours, 2),
                        "status": overtime_status,
                        "reason": getattr(overtime, 'reason', '') or "No reason provided"
                    })

            return Response({
                "name": employee.name,
                "designation": employee.designation.desig_name if employee.designation else "N/A",
                "department": getattr(employee.user, "department", "N/A"),  # Safe access
                "total_overtime": total_overtime,
                "overtime_history": filtered_overtime
            }, status=200)
        
        except Exception as e:
            return Response({"error": f"Internal Server Error: {str(e)}"}, status=500)

class FirstAssignOvertimeView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        
        data = request.data
        employee_id = data.get("employee_id")
        date_str = data.get("date")
        hours = data.get("hours")
        reason = data.get("reason", "").strip()  # Get reason with empty string as default

        
        if not employee_id or not date_str or not hours:
            return Response({"error": "Employee ID, date, and hours are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            overtime_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            today = date.today()
            if overtime_date <= today:  # Blocks past and present dates
                return Response({"error": "Overtime can only be assigned for future dates."}, status=400)

        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        employee = get_object_or_404(Employee, id=employee_id)

        if Overtime.objects.filter(employee=employee, date=overtime_date).exists():
            return Response(
                {"error": f"Overtime already exists for {employee.name} on {date_str}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            hours = float(hours)
            if hours <= 0:
                return Response({"error": "Hours must be positive."}, status=400)
            if hours > 12:  # Example: Max 12 hours per day
                return Response({"error": "Overtime cannot exceed 12 hours per day."}, status=400)
        except ValueError:
            return Response({"error": "Invalid hours value."}, status=400)

        if not reason:
            return Response({"error": "Reason for overtime is required."}, status=status.HTTP_400_BAD_REQUEST)

        overtime = Overtime.objects.create(
            employee=employee,
            date=overtime_date,
            hours=hours,
            reason=reason
        )

        return Response({
            "message": "Overtime assigned successfully",
            "overtime": OvertimeSerializer(overtime).data
        }, status=status.HTTP_201_CREATED)

############### Attendance  ##################
from django.db.models import ExpressionWrapper, DurationField
class HRAttendanceView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get date filters from query params
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            month_str = request.query_params.get('month')  # Accepts "YYYY-MM"

            # Default date range (last 30 days)
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)

            if month_str:  # Handle Flutter's "YYYY-MM" input
                try:
                    year, month = map(int, month_str.split('-'))
                    start_date = date(year, month, 1)
                    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)  # Last day of month
                except ValueError:
                    return Response({"error": "Invalid month format. Use 'YYYY-MM'."}, 
                                  status=status.HTTP_400_BAD_REQUEST)

            elif start_date_str and end_date_str:  # Handle React's "YYYY-MM-DD"
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except ValueError:
                    return Response({"error": "Invalid date format. Use 'YYYY-MM-DD'."},
                                  status=status.HTTP_400_BAD_REQUEST)

            # Determine user role and filter accordingly
            user_role = request.user.role.lower()

            if user_role == "hr":
                employee_filter = {"user__role": "employee"}  # HR sees only employees
            elif user_role == "admin":
                employee_filter = {}  # Admin sees all users
            else:
                return Response(
                    {"error": "Access denied! Only HR and Admin can view attendance reports."},
                    status=status.HTTP_403_FORBIDDEN
                )

            # 1. Employee Statistics
            total_employees = Employee.objects.filter(**employee_filter).count()

            # 2. Attendance Statistics
            attendance_stats = Attendance.objects.filter(
                date__range=[start_date, end_date], 
                employee__in=Employee.objects.filter(**employee_filter)
            ).aggregate(
                total_present=Count('id', filter=Q(status='present')),
                total_absent=Count('id', filter=Q(status='absent')),
                total_late=Count('id', filter=Q(status='late'))
            )

            # Calculate attendance percentage
            if total_employees > 0:
                total_days = total_employees * (end_date - start_date).days
                present_days = attendance_stats['total_present'] or 0
                attendance_percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0
            else:
                attendance_percentage = 0

            # 3. Pending Requests
            pending_attendance = AttendanceRequest.objects.filter(status='pending').count()
            pending_leaves = LeaveRequest.objects.filter(status='Pending').count()

            # 4. Detailed Employee Report with FIXED overtime calculation
            employees = Employee.objects.filter(**employee_filter).select_related(
                'designation', 'community', 'user'
            ).annotate(
                work_days = Count(
                    'attendance',
                    filter=Q(attendance__status__in=['present', 'late']) & Q(attendance__date__range=[start_date, end_date]),
                    distinct=True
                ),

                absent_days=Count(
                    'attendance',
                    filter=Q(attendance__status='absent') & Q(attendance__date__range=[start_date, end_date]),
                    distinct=True
                ),
                approved_leaves=Count(
                    'leaverequest',
                    filter=(
                        Q(leaverequest__status='Approved') | Q(leaverequest__status='Cancel Rejected')
                    ) & Q(leaverequest__start_date__lte=end_date) & Q(leaverequest__end_date__gte=start_date),
                    distinct=True
                ),
                total_overtime=Sum(
                    ExpressionWrapper(
                        F('attendance__check_out') - F('attendance__check_in') - timedelta(hours=8) ,
                        output_field=DurationField()
                    ),
                    filter=Q(attendance__date__range=[start_date, end_date]),
                    distinct=True
                )
            )

            employee_data = []
            for emp in employees:
                # Format overtime hours properly
                if emp.total_overtime:
    # Subtract 8 hours per working day to get actual overtime
                    total_seconds = emp.total_overtime.total_seconds()
                    expected_seconds = emp.work_days * 8 * 3600
                    overtime_seconds = total_seconds - expected_seconds

                    overtime_seconds = max(overtime_seconds, 0)  # Prevent negatives

                    hours = int(overtime_seconds // 3600)
                    minutes = int((overtime_seconds % 3600) // 60)
                    overtime_str = f"{hours}:{minutes:02d}"
                else:
                    overtime_str = "0:00"

                employee_data.append({
                    'emp_id': emp.id,
                    'emp_num': emp.emp_num,
                    'name': emp.name,
                    'designation': emp.designation.desig_name if emp.designation else None,
                    'community': emp.community.community_name if emp.community else None,
                    'work_days': emp.work_days or 0,
                    'absent_days': emp.absent_days or 0,
                    'approved_leaves': emp.approved_leaves or 0,
                    'total_overtime': overtime_str,
                    'image': request.build_absolute_uri(emp.image.url) if emp.image else None
                })

            # 5. Generate Report Data
            report_data = {
                'date_range': {
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d')
                },
                'summary': {
                    'total_employees': total_employees,
                    'attendance_percentage': attendance_percentage,
                    'present_days': attendance_stats['total_present'] or 0,
                    'absent_days': attendance_stats['total_absent'] or 0,
                    'late_days': attendance_stats['total_late'] or 0,
                    'pending_attendance_requests': pending_attendance,
                    'pending_leave_requests': pending_leaves
                },
                'employees': employee_data
            }

            return Response(report_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class HRAttendanceRequestView(APIView):
    def get(self, request):
        """Fetch all pending attendance requests submitted by employees."""
        try:
            pending_requests = AttendanceRequest.objects.filter(status='pending').select_related('employee')

            requests_data = []
            for req in pending_requests:
                requests_data.append({
                    "id": req.id,
                    "employee_id": req.employee.id,
                    "employee_name": req.employee.name,
                    "date": req.date.strftime('%Y-%m-%d'),
                    "work_type": req.work_type,
                    "location": req.location,
                    "check_in": req.check_in.strftime('%H:%M') if req.check_in else None,
                    "check_out": req.check_out.strftime('%H:%M') if req.check_out else None,
                    "image": request.build_absolute_uri(req.image.url) if req.image else None,  
                })

            return Response({"pending_attendance_requests": requests_data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        """Approve or Reject an attendance request."""
        try:
            request_id = request.data.get("id")
            action = request.data.get("action")  # "approve" or "reject"

            if not request_id or action not in ["approve", "reject"]:
                return Response(
                    {"error": "Invalid request. Provide 'id' and 'action' (approve/reject)."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            attendance_request = get_object_or_404(AttendanceRequest, id=request_id)

            if action == "approve":
                attendance_request.status = "approved"

                # ✅ Automatically create an Attendance entry if approved
                Attendance.objects.create(
                    employee=attendance_request.employee,
                    date=attendance_request.date,
                    check_in=attendance_request.check_in,
                    check_out=attendance_request.check_out,
                    status="present"
                )

            elif action == "reject":
                attendance_request.status = "rejected"

            attendance_request.save()

            return Response(
                {"message": f"Attendance request {action}d successfully."},
                status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GenerateReportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get response format and filters
            response_format = request.query_params.get('format', 'json')
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            department_id = request.query_params.get('department')

            # Set default date range (last 30 days) if not provided
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=30)

            if start_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

            # Base employee query
            employees = Employee.objects.filter(user__role='employee').select_related(
                'designation', 'community', 'user'
            )

            # Apply department filter if provided
            if department_id and department_id != 'All':
                employees = employees.filter(community_id=department_id)

            total_employees = employees.count()

            # Global stats
            attendance_stats = Attendance.objects.filter(
                date__range=[start_date, end_date],
                employee__in=employees
            ).aggregate(
                total_present=Count('id', filter=Q(status='present')),
                total_absent=Count('id', filter=Q(status='absent')),
                total_late=Count('id', filter=Q(status='late'))
            )

            total_days = total_employees * ((end_date - start_date).days + 1)
            present_days = attendance_stats['total_present'] or 0
            attendance_percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0

            pending_attendance = AttendanceRequest.objects.filter(status='pending').count()
            pending_leaves = LeaveRequest.objects.filter(status='Pending').count()

            # Get all departments for the filter dropdown
            departments = list(Community.objects.all().values('id', 'community_name'))

            # Annotate basic stats per employee
            employees = employees.annotate(
                present_days=Count(
                    'attendance',
                    filter=Q(attendance__status='present', attendance__date__range=[start_date, end_date])
                ),
                absent_days=Count(
                    'attendance',
                    filter=Q(attendance__status='absent', attendance__date__range=[start_date, end_date])
                ),
                late_days=Count(
                    'attendance',
                    filter=Q(attendance__status='late', attendance__date__range=[start_date, end_date])
                ),
                approved_leaves=Count(
                    'leaverequest',
                    filter=Q(
                        leaverequest__status='Approved',
                        leaverequest__start_date__lte=end_date,
                        leaverequest__end_date__gte=start_date
                    )
                )
            )

            # Prepare employee data with overtime calculation
            employee_data = []
            for emp in employees:
                total_overtime_seconds = 0
                attendance_qs = Attendance.objects.filter(
                    employee=emp, 
                    date__range=[start_date, end_date]
                )

                for att in attendance_qs:
                    if att.check_in and att.check_out:
                        work_duration = (
                            datetime.combine(att.date, att.check_out) - 
                            datetime.combine(att.date, att.check_in)
                        ).total_seconds()
                        overtime = max(0, work_duration - 8 * 3600)  # over 8 hrs only
                        total_overtime_seconds += overtime

                # Format overtime as HH:MM:SS
                total_overtime = str(timedelta(seconds=total_overtime_seconds))
                if total_overtime == '0:00:00':
                    total_overtime = '0:00'

                employee_data.append({
                    'id': emp.id,
                    'emp_num': emp.emp_num,
                    'name': emp.name,
                    'designation': {'desig_name': emp.designation.desig_name} if emp.designation else None,
                    'community': {'community_name': emp.community.community_name} if emp.community else None,
                    'present_days': emp.present_days or 0,
                    'absent_days': emp.absent_days or 0,
                    'late_days': emp.late_days or 0,
                    'approved_leaves': emp.approved_leaves or 0,
                    'total_overtime': total_overtime
                })

            # Return in requested format
            if response_format == 'excel':
                return self.generate_excel_report(
                    start_date, end_date, total_employees, attendance_stats,
                    attendance_percentage, pending_attendance, pending_leaves, employee_data
                )
            else:
                return Response({
                    'status': 'success',
                    'data': {
                        'employee_data': employee_data,
                        'total_employees': total_employees,
                        'attendance_percentage': attendance_percentage,
                        'attendance_stats': {
                            'total_present': attendance_stats['total_present'] or 0,
                            'total_absent': attendance_stats['total_absent'] or 0,
                            'total_late': attendance_stats['total_late'] or 0
                        },
                        'pending_attendance': pending_attendance,
                        'pending_leaves': pending_leaves,
                        'departments': departments,
                        'start_date': start_date.strftime('%Y-%m-%d'),
                        'end_date': end_date.strftime('%Y-%m-%d')
                    }
                })

        except Exception as e:
            return Response({'status': 'error', 'message': str(e)}, status=500)

    def generate_excel_report(self, start_date, end_date, total_employees, attendance_stats,
                            attendance_percentage, pending_attendance, pending_leaves, employees):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR Attendance Report"

        # Add headers
        headers = ["Emp ID", "Emp Num", "Name", "Designation", "Community",
                  "Present Days", "Absent Days", "Late Days", "Approved Leaves", "Total Overtime"]
        ws.append(headers)
        
        # Style headers
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        # Add data rows
        for emp in employees:
            ws.append([
                emp['id'],
                emp['emp_num'],
                emp['name'],
                emp['designation']['desig_name'] if emp['designation'] else "N/A",
                emp['community']['community_name'] if emp['community'] else "N/A",
                emp['present_days'],
                emp['absent_days'],
                emp['late_days'],
                emp['approved_leaves'],
                emp['total_overtime']
            ])

        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            status=200
        )
        response['Content-Disposition'] = f'attachment; filename=HR_Report_{start_date}_to_{end_date}.xlsx'
        wb.save(response)
        return response


from datetime import datetime, timedelta

class EmployeeAttendanceDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
    
        employee = get_object_or_404(Employee, id=employee_id)

        emp_details = {
            "name": employee.name,
            "designation": getattr(employee.designation, "desig_name", "N/A"),
            "department": employee.user.department if isinstance(employee.user.department, str) else "N/A",
            "emp_num": employee.emp_num,
        }

        #  Fetch unpaid leave count
        unpaid_leaves = LeaveBalance.objects.filter(employee=employee, leave_policy__leave_type="unpaid").count()

        #  Fetch total overtime
        total_overtime = (
            Overtime.objects.filter(employee=employee)
            .aggregate(Sum('hours'))['hours__sum'] or 0
        )

        # Fetch attendance history
        attendance_records = Attendance.objects.filter(employee=employee).order_by('-date')

        attendance_data = []
        attendance_summary = {"present": 0, "late": 0, "absent": 0, "total": 0}

        for record in attendance_records:
            status_key = record.status.lower()

            # Convert time to datetime for subtraction
            if record.check_in and record.check_out:
                check_in_dt = datetime.combine(record.date, record.check_in)
                check_out_dt = datetime.combine(record.date, record.check_out)
                work_hours = check_out_dt - check_in_dt
                total_hours = str(work_hours).split('.')[0]  # Convert timedelta to HH:MM
            else:
                total_hours = "0:00"  # No check-in or check-out

            attendance_data.append({
                "date": record.date.strftime("%d-%m-%Y"),
                "check_in": record.check_in.strftime("%I:%M%p") if record.check_in else "-",
                "check_out": record.check_out.strftime("%I:%M%p") if record.check_out else "-",
                "overtime": f"{record.overtime_hours} hrs" if hasattr(record, "overtime_hours") and record.overtime_hours else "-",
                "status": record.status.capitalize(),
                "total_hours": total_hours,  # Fixed working hours calculation
            })

            # Update attendance summary
            if status_key in attendance_summary:
                attendance_summary[status_key] += 1
            attendance_summary["total"] += 1

        return Response({
            "employee_details": emp_details,
            "unpaid_leaves": unpaid_leaves,
            "total_overtime": total_overtime,
            "attendance_records": attendance_data,
            "attendance_summary": attendance_summary,  # For graphical statistics
        }, status=status.HTTP_200_OK)

class AddAttendanceRecordView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, employee_id):
        try:
            # Extract data from request
            date_str = request.data.get("date")  # Expected format: "YYYY-MM-DD"
            check_in_str = request.data.get("check_in")  # Expected format: "HH:MM"
            check_out_str = request.data.get("check_out")  # Expected format: "HH:MM"

            if not date_str or not check_in_str or not check_out_str:
                return Response({"error": "Missing required fields (date, check_in, check_out)"}, status=400)

            # Convert date & time to datetime objects
            attendance_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            check_in = datetime.strptime(check_in_str, "%H:%M").time()
            check_out = datetime.strptime(check_out_str, "%H:%M").time()

            # Validate check-in and check-out order
            if check_out <= check_in:
                return Response({"error": "Check-out time must be later than check-in time"}, status=400)

            # Fetch employee based on URL parameter
            employee = get_object_or_404(Employee, id=employee_id)

            # Calculate worked hours
            worked_seconds = (datetime.combine(attendance_date, check_out) - datetime.combine(attendance_date, check_in)).seconds
            worked_hours = round(worked_seconds / 3600, 2)  # Convert seconds to hours

            # Create attendance record
            attendance = Attendance.objects.create(
                employee=employee,
                date=attendance_date,
                check_in=check_in,
                check_out=check_out,
                # worked_hours=worked_hours,
                status="present",  # Default status
            )

            return Response({"message": "Attendance recorded successfully!", "attendance_id": attendance.id}, status=201)

        except ValueError as ve:
            return Response({"error": f"Invalid date or time format: {str(ve)}"}, status=400)

        except Exception as e:
            return Response({"error": str(e)}, status=400)

class EmployeeAttendanceDashDetailView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Fetch the logged-in employee
        employee = get_object_or_404(Employee, user=request.user)

        # Fetch employee number, designation, department
        emp_details = {
            "name": employee.name,
            "designation": getattr(employee.designation, "desig_name", "N/A"),
            "department": employee.user.department if isinstance(employee.user.department, str) else "N/A",
            "emp_num": employee.emp_num,
        }

        # Fetch unpaid leave count
        unpaid_leaves = LeaveBalance.objects.filter(employee=employee, leave_policy__leave_type="unpaid").count()

        # Fetch total overtime
        total_overtime = (
            Overtime.objects.filter(employee=employee)
            .aggregate(Sum('hours'))['hours__sum'] or 0
        )

        # Fetch attendance history of the logged-in employee
        attendance_records = Attendance.objects.filter(employee=employee).order_by('-date')

        attendance_data = []
        attendance_summary = {"present": 0, "late": 0, "absent": 0, "total": 0}

        for record in attendance_records:
            status_key = record.status.lower()

            #Convert time to datetime for subtraction
            if record.check_in and record.check_out:
                check_in_dt = datetime.combine(record.date, record.check_in)
                check_out_dt = datetime.combine(record.date, record.check_out)
                work_hours = check_out_dt - check_in_dt
                total_hours = str(work_hours).split('.')[0]  # Convert timedelta to HH:MM
            else:
                total_hours = "0:00"  # No check-in or check-out

            attendance_data.append({
                "date": record.date.strftime("%d-%m-%Y"),
                "check_in": record.check_in.strftime("%I:%M%p") if record.check_in else "-",
                "check_out": record.check_out.strftime("%I:%M%p") if record.check_out else "-",
                "overtime": f"{record.overtime_hours} hrs" if hasattr(record, "overtime_hours") and record.overtime_hours else "-",
                "status": record.status.capitalize(),
                "total_hours": total_hours,  #Fixed working hours calculation
            })

            #Update attendance summary
            if status_key in attendance_summary:
                attendance_summary[status_key] += 1
            attendance_summary["total"] += 1

        return Response({
            "employee_details": emp_details,
            "unpaid_leaves": unpaid_leaves,
            "total_overtime": total_overtime,
            "attendance_records": attendance_data,
            "attendance_summary": attendance_summary,  # For graphical statistics
        }, status=status.HTTP_200_OK)

class AdminAttendanceReportView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Ensure only Admins can access
            if not hasattr(request.user, 'role') or request.user.role.lower() != "admin":
                return Response({"error": "Access Denied! Only Admins can generate reports."},
                                status=status.HTTP_403_FORBIDDEN)

            # Get date filters from request
            start_date = request.query_params.get('fromDate')
            end_date = request.query_params.get('toDate')

            if not start_date or not end_date:
                return Response({"error": "Please provide fromDate and toDate in YYYY-MM-DD format."},
                                status=status.HTTP_400_BAD_REQUEST)

            # Convert string dates to Date objects
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

            # Fetch employees with proper overtime calculation
            employees = Employee.objects.all().annotate(
                work_days=Count('attendance', 
                              filter=Q(attendance__status='present', 
                                      attendance__date__range=[start_date, end_date]), 
                              distinct=True),
                leave_taken=Count('leaverequest', 
                                filter=Q(leaverequest__status='Approved', 
                                        leaverequest__start_date__lte=end_date, 
                                        leaverequest__end_date__gte=start_date), 
                                distinct=True),
                # Correct overtime calculation
                total_overtime=Sum(
                    ExpressionWrapper(
                        F('attendance__check_out') - F('attendance__check_in') - timedelta(hours=8),
                        output_field=DurationField()
                    ),
                    filter=Q(attendance__date__range=[start_date, end_date],
                            attendance__status='present'),
                    distinct=True
                )
            )

            # Prepare data for JSON response
            report_data = []
            for emp in employees:
                # Format overtime properly
                overtime_hours = 0
                if emp.total_overtime:
                    total_seconds = emp.total_overtime.total_seconds()
                    if total_seconds > 0:  # Only count positive overtime
                        overtime_hours = round(total_seconds / 3600, 2)

                report_data.append({
                    'name': emp.name,
                    'workDays': emp.work_days or 0,
                    'leaves': emp.leave_taken or 0,
                    'overtime': overtime_hours
                })

            return Response({
                "message": "Report generated successfully!",
                "data": report_data,
                "download_url": f"/api/admin/attendance-report/download/?fromDate={start_date}&toDate={end_date}"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AdminAttendanceReportDownloadView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            if request.user.role.lower() != "admin":
                return Response({"error": "Access Denied! Only Admins can download reports."},
                                status=status.HTTP_403_FORBIDDEN)

            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            if not start_date_str or not end_date_str:
                return Response({"error": "Please provide start_date and end_date in YYYY-MM-DD format."},
                                status=status.HTTP_400_BAD_REQUEST)

            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

            employees = Employee.objects.all().annotate(
                work_days=Count('attendance', filter=Q(attendance__status='present', attendance__date__range=[start_date, end_date])),
                leave_taken=Count('leaverequest', filter=Q(leaverequest__status='Approved', leaverequest__start_date__lte=end_date, leaverequest__end_date__gte=start_date)),
                total_overtime=Sum('overtime__hours', filter=Q(overtime__date__range=[start_date, end_date], overtime__status='completed'))
            )

            # Create DataFrame for Excel
            data = [{
                'Employee ID': emp.id,
                'Name': emp.name,
                'Work Days': emp.work_days or 0,
                'Leave Taken': emp.leave_taken or 0,
                'Total Overtime (Hours)': emp.total_overtime or 0
            } for emp in employees]

            df = pd.DataFrame(data)

            # Generate Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Attendance_Report_{start_date}_{end_date}.xlsx"'
            df.to_excel(response, index=False)

            return response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)